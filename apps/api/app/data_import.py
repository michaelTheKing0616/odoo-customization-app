"""Bulk CSV/XLSX → Odoo record create/upsert via public RPC."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any, Literal

from odoo_client import OdooClient
from odoo_client.client import OdooClientError

Mode = Literal["create", "upsert"]


@dataclass
class RowResult:
    row_index: int  # 1-based data row (excluding header)
    ok: bool
    record_id: int | None = None
    action: str | None = None  # create | write | skip
    error: str | None = None
    values: dict[str, Any] = field(default_factory=dict)


@dataclass
class ImportPreview:
    headers: list[str]
    sample_rows: list[dict[str, str]]
    row_count: int
    suggested_model: str | None
    field_hints: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportCommitResult:
    ok: bool
    created: int = 0
    updated: int = 0
    failed: int = 0
    skipped: int = 0
    results: list[RowResult] = field(default_factory=list)
    message: str = ""


_HEADER_ALIASES: dict[str, dict[str, str]] = {
    "res.partner": {
        "name": "name",
        "display_name": "name",
        "email": "email",
        "phone": "phone",
        "mobile": "mobile",
        "street": "street",
        "city": "city",
        "zip": "zip",
        "vat": "vat",
        "is_company": "is_company",
        "company_type": "company_type",
        "comment": "comment",
        "website": "website",
    },
    "product.template": {
        "name": "name",
        "default_code": "default_code",
        "list_price": "list_price",
        "standard_price": "standard_price",
        "type": "type",
        "barcode": "barcode",
        "description": "description",
        "description_sale": "description_sale",
    },
    "product.product": {
        "name": "name",
        "default_code": "default_code",
        "list_price": "list_price",
        "barcode": "barcode",
    },
}


def parse_tabular(raw: bytes, filename: str) -> tuple[list[str], list[dict[str, str]]]:
    """Return (headers, rows as dicts of string cells)."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(raw)
    # default CSV (also .txt)
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV has no header row")
    headers = [str(h).strip() for h in reader.fieldnames if h is not None and str(h).strip()]
    rows: list[dict[str, str]] = []
    for row in reader:
        cleaned = {
            str(k).strip(): ("" if v is None else str(v).strip())
            for k, v in row.items()
            if k is not None and str(k).strip()
        }
        if any(cleaned.values()):
            rows.append(cleaned)
    return headers, rows


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "XLSX support requires openpyxl — install API deps (openpyxl)"
        ) from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("XLSX has no active sheet")
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration as exc:
        raise ValueError("XLSX sheet is empty") from exc
    headers = [
        str(c).strip()
        for c in header_row
        if c is not None and str(c).strip()
    ]
    if not headers:
        raise ValueError("XLSX has no header row")
    rows: list[dict[str, str]] = []
    for raw_row in it:
        cells = list(raw_row or [])
        cleaned: dict[str, str] = {}
        for i, h in enumerate(headers):
            val = cells[i] if i < len(cells) else None
            cleaned[h] = "" if val is None else str(val).strip()
        if any(cleaned.values()):
            rows.append(cleaned)
    return headers, rows


def suggest_mapping(model: str, headers: list[str]) -> dict[str, str]:
    """Map CSV header → Odoo field name (best-effort)."""
    aliases = _HEADER_ALIASES.get(model, {})
    out: dict[str, str] = {}
    for h in headers:
        key = h.strip()
        low = key.lower().replace(" ", "_")
        if low in aliases:
            out[key] = aliases[low]
            continue
        # x_ fields or exact match
        if re.match(r"^x_[a-z0-9_]+$", low):
            out[key] = low
        elif low in {"id", "external_id", "xml_id"}:
            out[key] = "__external_id__"
        else:
            out[key] = low  # hope field exists; validate later
    return out


def build_preview(
    *,
    headers: list[str],
    rows: list[dict[str, str]],
    model: str | None,
) -> ImportPreview:
    suggested = model
    if not suggested:
        lows = {h.lower() for h in headers}
        if "email" in lows or "vat" in lows:
            suggested = "res.partner"
        elif "list_price" in lows or "default_code" in lows:
            suggested = "product.template"
        elif any(h.startswith("x_") for h in lows):
            suggested = None
        else:
            suggested = "res.partner"
    mapping = suggest_mapping(suggested or "res.partner", headers) if suggested else {}
    hints = [
        {"header": h, "suggested_field": mapping.get(h, ""), "sample": rows[0].get(h, "") if rows else ""}
        for h in headers
    ]
    return ImportPreview(
        headers=headers,
        sample_rows=rows[:5],
        row_count=len(rows),
        suggested_model=suggested,
        field_hints=hints,
        warnings=[],
    )


def _coerce_value(raw: str, ttype: str) -> Any:
    if raw == "":
        return False if ttype == "boolean" else None
    if ttype == "boolean":
        return raw.lower() in {"1", "true", "yes", "y", "t"}
    if ttype in {"integer", "many2one"}:
        try:
            return int(float(raw))
        except ValueError:
            return raw  # name resolution later for m2o
    if ttype in {"float", "monetary"}:
        return float(raw.replace(",", ""))
    return raw


def _field_meta(client: OdooClient, model: str) -> dict[str, dict[str, Any]]:
    rows = client.execute_kw(
        "ir.model.fields",
        "search_read",
        [[("model", "=", model)]],
        {
            "fields": ["name", "ttype", "relation", "required", "readonly"],
            "limit": 5000,
        },
    )
    return {str(r["name"]): r for r in rows}


def _resolve_m2o(
    client: OdooClient,
    relation: str,
    value: Any,
) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    # external id module.name
    if "." in text and " " not in text:
        ids = client.execute_kw(
            "ir.model.data",
            "search",
            [[("name", "=", text.split(".", 1)[1]), ("module", "=", text.split(".", 1)[0])]],
            {"limit": 1},
        )
        if ids:
            data = client.execute_kw(
                "ir.model.data", "read", [ids], {"fields": ["res_id"]}
            )
            return int(data[0]["res_id"])
    # name search
    found = client.execute_kw(
        relation,
        "name_search",
        [text],
        {"limit": 2},
    )
    if not found:
        raise ValueError(f"No {relation} matching {text!r}")
    if len(found) > 1:
        raise ValueError(f"Ambiguous {relation} for {text!r} ({len(found)} matches)")
    return int(found[0][0])


def _row_to_vals(
    client: OdooClient,
    model: str,
    row: dict[str, str],
    mapping: dict[str, str],
    meta: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], str | None]:
    """Return (vals, external_id_or_none)."""
    vals: dict[str, Any] = {}
    external_id: str | None = None
    for header, field_name in mapping.items():
        if not field_name or field_name.startswith("__skip"):
            continue
        raw = row.get(header, "")
        if field_name == "__external_id__":
            external_id = raw or None
            continue
        info = meta.get(field_name)
        if info is None:
            raise ValueError(f"Unknown field {field_name} on {model}")
        if info.get("readonly") and field_name not in {"id"}:
            continue
        ttype = str(info.get("ttype") or "char")
        if raw == "" and not info.get("required"):
            continue
        coerced = _coerce_value(raw, ttype)
        if ttype == "many2one":
            rel = info.get("relation")
            if not rel:
                raise ValueError(f"{field_name} has no relation")
            coerced = _resolve_m2o(client, str(rel), coerced if coerced is not None else raw)
        if coerced is None and info.get("required"):
            raise ValueError(f"Required field {field_name} is empty")
        if coerced is not None:
            vals[field_name] = coerced
    return vals, external_id


def _find_by_external_id(client: OdooClient, model: str, xml_id: str) -> int | None:
    if "." not in xml_id:
        return None
    module, name = xml_id.split(".", 1)
    ids = client.execute_kw(
        "ir.model.data",
        "search_read",
        [[("module", "=", module), ("name", "=", name), ("model", "=", model)]],
        {"fields": ["res_id"], "limit": 1},
    )
    if not ids:
        return None
    return int(ids[0]["res_id"])


def _find_by_keys(
    client: OdooClient,
    model: str,
    vals: dict[str, Any],
    match_fields: list[str],
) -> int | None:
    domain: list[Any] = []
    for f in match_fields:
        if f not in vals:
            return None
        domain.append((f, "=", vals[f]))
    if not domain:
        return None
    ids = client.execute_kw(model, "search", [domain], {"limit": 2})
    if not ids:
        return None
    if len(ids) > 1:
        raise ValueError(f"Upsert match ambiguous for {match_fields}: {ids}")
    return int(ids[0])


def dry_run_or_commit(
    client: OdooClient,
    *,
    model: str,
    rows: list[dict[str, str]],
    mapping: dict[str, str],
    mode: Mode = "create",
    match_fields: list[str] | None = None,
    dry_run: bool = True,
    batch_size: int = 50,
    rpc_context: dict[str, Any] | None = None,
) -> ImportCommitResult:
    if not client.model_exists(model):
        raise OdooClientError(f"Model {model} not found on this Odoo")
    if batch_size < 1 or batch_size > 500:
        raise ValueError("batch_size must be between 1 and 500")
    meta = _field_meta(client, model)
    match_fields = match_fields or []
    results: list[RowResult] = []
    created = updated = failed = skipped = 0
    pending_creates: list[tuple[int, dict[str, Any]]] = []

    def _kw(extra: dict[str, Any] | None = None) -> dict[str, Any]:
        out = dict(extra or {})
        if rpc_context:
            out["context"] = {**(out.get("context") or {}), **rpc_context}
        return out

    def flush_creates() -> None:
        nonlocal created, failed
        if not pending_creates:
            return
        for start in range(0, len(pending_creates), batch_size):
            chunk = pending_creates[start : start + batch_size]
            vals_list = [vals for _, vals in chunk]
            try:
                if dry_run:
                    for idx, vals in chunk:
                        created += 1
                        results.append(
                            RowResult(
                                row_index=idx,
                                ok=True,
                                action="create",
                                values=vals,
                            )
                        )
                    continue
                new_ids = client.execute_kw(model, "create", [vals_list], _kw())
                if isinstance(new_ids, int):
                    new_ids = [new_ids]
                if not isinstance(new_ids, list) or len(new_ids) != len(chunk):
                    raise OdooClientError(
                        f"Batch create returned unexpected ids: {new_ids!r}"
                    )
                for (idx, vals), new_id in zip(chunk, new_ids, strict=True):
                    created += 1
                    results.append(
                        RowResult(
                            row_index=idx,
                            ok=True,
                            record_id=int(new_id),
                            action="create",
                            values=vals,
                        )
                    )
            except Exception as exc:  # noqa: BLE001
                # Fall back to per-row so one bad row does not discard the batch.
                for idx, vals in chunk:
                    try:
                        if dry_run:
                            created += 1
                            results.append(
                                RowResult(
                                    row_index=idx,
                                    ok=True,
                                    action="create",
                                    values=vals,
                                )
                            )
                            continue
                        new_id = int(client.execute_kw(model, "create", [vals], _kw()))
                        created += 1
                        results.append(
                            RowResult(
                                row_index=idx,
                                ok=True,
                                record_id=new_id,
                                action="create",
                                values=vals,
                            )
                        )
                    except Exception as row_exc:  # noqa: BLE001
                        failed += 1
                        results.append(
                            RowResult(
                                row_index=idx,
                                ok=False,
                                error=str(row_exc),
                                values=vals,
                            )
                        )
                if not dry_run and "Batch create" not in str(exc):
                    # Keep first batch error only when all per-row also failed path used
                    pass
        pending_creates.clear()

    for idx, row in enumerate(rows, start=1):
        try:
            vals, external_id = _row_to_vals(client, model, row, mapping, meta)
            if not vals and not external_id:
                skipped += 1
                results.append(
                    RowResult(row_index=idx, ok=True, action="skip", error="empty row")
                )
                continue
            existing_id: int | None = None
            if mode == "upsert":
                if external_id:
                    existing_id = _find_by_external_id(client, model, external_id)
                if existing_id is None and match_fields:
                    existing_id = _find_by_keys(client, model, vals, match_fields)

            if existing_id:
                flush_creates()
                if dry_run:
                    updated += 1
                    results.append(
                        RowResult(
                            row_index=idx,
                            ok=True,
                            record_id=existing_id,
                            action="write",
                            values=vals,
                        )
                    )
                else:
                    client.execute_kw(model, "write", [[existing_id], vals], _kw())
                    updated += 1
                    results.append(
                        RowResult(
                            row_index=idx,
                            ok=True,
                            record_id=existing_id,
                            action="write",
                            values=vals,
                        )
                    )
            else:
                pending_creates.append((idx, vals))
                if len(pending_creates) >= batch_size:
                    flush_creates()
        except Exception as exc:  # noqa: BLE001
            failed += 1
            results.append(
                RowResult(row_index=idx, ok=False, error=str(exc), values=dict(row))
            )

    flush_creates()
    results.sort(key=lambda r: r.row_index)
    message = (
        f"{'Dry-run' if dry_run else 'Commit'}: "
        f"{created} create, {updated} update, {failed} failed, {skipped} skipped"
        f" (batch_size={batch_size})"
    )
    return ImportCommitResult(
        ok=failed == 0,
        created=created,
        updated=updated,
        failed=failed,
        skipped=skipped,
        results=results,
        message=message,
    )


def template_csv(model: str) -> str:
    """Return a starter CSV string for common models / industry seeds."""
    from app.industry_seeds import template_csv_for_model

    seeded = template_csv_for_model(model)
    if seeded:
        return seeded
    if model == "res.partner":
        return (
            "name,email,phone,is_company,street,city\n"
            "Acme Corp,info@acme.example,+10000000000,true,1 Main St,Lagos\n"
        )
    if model in {"product.template", "product.product"}:
        return (
            "name,default_code,list_price,type\n"
            "Sample Product,SKU-001,19.99,consu\n"
        )
    return "x_name,x_notes\nExample row,notes here\n"


def results_to_error_csv(results: list[RowResult]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["row_index", "ok", "action", "record_id", "error"])
    for r in results:
        if r.ok and not r.error:
            continue
        w.writerow([r.row_index, r.ok, r.action or "", r.record_id or "", r.error or ""])
    return buf.getvalue()
